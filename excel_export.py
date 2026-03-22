"""
Módulo 7 — Exportación a Excel profesional
=============================================
Genera un archivo .xlsx con hojas separadas:
  1. Resumen del perfil del inversor
  2. Análisis de acciones individuales
  3. Portafolios candidatos (P1-P5) con métricas
  4. Matrices AHP (criterios + alternativas)
  5. Ranking final con gráfico
  6. Composición recomendada del portafolio ganador

Usa openpyxl para formateo profesional con colores,
bordes y gráficos tipo financiero.

Autor: [Tu nombre] — TFG
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from typing import Dict, Optional
from profiles import CRITERIA, CRITERIA_ORDER, PROFILE_WEIGHTS


# ─────────────────────────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────────────────────────

# Colores corporativos
BLUE_DARK = "1F3864"
BLUE_MED = "2E5090"
BLUE_LIGHT = "D6E4F0"
GRAY_LIGHT = "F2F2F2"
GREEN = "548235"
RED = "C00000"
WHITE = "FFFFFF"

# Fuentes
FONT_TITLE = Font(name="Arial", size=16, bold=True, color=BLUE_DARK)
FONT_HEADER = Font(name="Arial", size=11, bold=True, color=WHITE)
FONT_SUBHEADER = Font(name="Arial", size=11, bold=True, color=BLUE_DARK)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_SMALL = Font(name="Arial", size=9, color="666666")
FONT_NUMBER = Font(name="Arial", size=10)
FONT_WINNER = Font(name="Arial", size=12, bold=True, color=GREEN)

# Rellenos
FILL_HEADER = PatternFill(start_color=BLUE_MED, end_color=BLUE_MED, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")
FILL_GRAY = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")
FILL_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_RED = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

# Bordes
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row, max_col):
    """Aplica estilo de cabecera a una fila."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, is_alt=False):
    """Aplica estilo a una celda de datos."""
    cell = ws.cell(row=row, column=col)
    cell.font = FONT_NORMAL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if is_alt:
        cell.fill = FILL_GRAY


def auto_width(ws, min_width=10, max_width=25):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# ─────────────────────────────────────────────────────────────────
# GENERACIÓN DE HOJAS
# ─────────────────────────────────────────────────────────────────

def create_profile_sheet(wb: Workbook, profile: str, filters):
    """Hoja 1: Resumen del perfil del inversor."""
    ws = wb.create_sheet("1. Perfil del inversor")

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Perfil: {profile.upper().replace('_', ' ')}"
    ws["A1"].font = FONT_TITLE

    ws["A3"] = "Configuración del perfil"
    ws["A3"].font = FONT_SUBHEADER

    # Info de filtros
    info = [
        ("Capitalización mínima", f"${filters.min_market_cap_bn:.0f}B"),
        ("Volumen medio mínimo", f"{filters.min_avg_volume:,.0f}"),
        ("Acciones por índice", str(filters.max_stocks_per_index)),
        ("N acciones finales", str(filters.final_n_stocks)),
        ("Descripción", filters.description),
    ]
    if filters.sectors_gics:
        info.append(("Sectores GICS", ", ".join(filters.sectors_gics)))
    if filters.min_dividend_yield:
        info.append(("Dividend yield mín.", f"{filters.min_dividend_yield}%"))
    if filters.min_esg_percentile:
        info.append(("ESG percentil mín.", f"{filters.min_esg_percentile}%"))

    for i, (label, value) in enumerate(info, 5):
        ws.cell(row=i, column=1, value=label).font = FONT_NORMAL
        ws.cell(row=i, column=1).fill = FILL_LIGHT
        ws.cell(row=i, column=2, value=value).font = FONT_NORMAL
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2).border = THIN_BORDER

    # Tabla de pesos AHP
    row = len(info) + 7
    ws.cell(row=row, column=1, value="Pesos AHP por criterio").font = FONT_SUBHEADER
    row += 2

    headers = ["Categoría", "Criterio", "Dirección", "Peso (1-9)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    weights = PROFILE_WEIGHTS[profile]
    for i, crit in enumerate(CRITERIA_ORDER):
        is_alt = i % 2 == 1
        ws.cell(row=row, column=1, value=CRITERIA[crit]["cat"].capitalize())
        ws.cell(row=row, column=2, value=CRITERIA[crit]["label"])
        ws.cell(row=row, column=3, value=CRITERIA[crit]["dir"].upper())
        ws.cell(row=row, column=4, value=weights[crit])
        for c in range(1, 5):
            style_data_cell(ws, row, c, is_alt)
        # Color condicional del peso
        peso = weights[crit]
        cell = ws.cell(row=row, column=4)
        if peso >= 7:
            cell.fill = FILL_GREEN
            cell.font = Font(name="Arial", size=10, bold=True, color=GREEN)
        elif peso <= 3:
            cell.fill = FILL_RED
            cell.font = Font(name="Arial", size=10, color=RED)
        row += 1

    auto_width(ws)


def create_portfolios_sheet(wb: Workbook, portfolios: pd.DataFrame):
    """Hoja 3: Portafolios candidatos con métricas."""
    ws = wb.create_sheet("3. Portafolios candidatos")

    ws.merge_cells("A1:F1")
    ws["A1"] = "Portafolios candidatos para análisis AHP"
    ws["A1"].font = FONT_TITLE

    display_cols = [
        ("nombre", "Portafolio"),
        ("rentabilidad", "Rentabilidad"),
        ("sharpe", "Sharpe"),
        ("sortino", "Sortino"),
        ("alpha", "Alpha"),
        ("volatilidad", "Volatilidad"),
        ("var_95", "VaR 95%"),
        ("cvar_95", "CVaR 95%"),
        ("max_drawdown", "Max Drawdown"),
        ("beta", "Beta"),
        ("tracking_error", "Track. Error"),
        ("rentab_kp", "Rentab-kp"),
        ("cv", "C.V."),
        ("skewness", "Asimetría"),
        ("corr_media", "Corr. media"),
        ("div_geo", "Div. geográf."),
    ]

    available_cols = [(key, label) for key, label in display_cols if key in portfolios.columns or key in portfolios.index.names]

    row = 3
    for c, (key, label) in enumerate(available_cols, 1):
        ws.cell(row=row, column=c, value=label)
    style_header_row(ws, row, len(available_cols))

    for i, (idx, prow) in enumerate(portfolios.iterrows()):
        row = 4 + i
        is_alt = i % 2 == 1
        for c, (key, label) in enumerate(available_cols, 1):
            if key == "nombre":
                val = idx if isinstance(idx, str) else prow.get("nombre", f"P{i+1}")
            else:
                val = prow.get(key, "")
            ws.cell(row=row, column=c, value=val)
            style_data_cell(ws, row, c, is_alt)
            # Formato porcentaje para métricas relevantes
            if key in ["rentabilidad", "volatilidad", "var_95", "cvar_95", "max_drawdown", "rentab_kp", "tracking_error"]:
                ws.cell(row=row, column=c).number_format = "0.00%"
            elif key in ["sharpe", "sortino", "alpha", "beta", "cv", "skewness", "corr_media", "div_geo"]:
                ws.cell(row=row, column=c).number_format = "0.00"

    auto_width(ws)


def create_ranking_sheet(wb: Workbook, ranking: pd.DataFrame, profile: str):
    """Hoja 5: Ranking final con gráfico."""
    ws = wb.create_sheet("5. Ranking AHP final")

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Ranking AHP — Perfil: {profile.upper().replace('_', ' ')}"
    ws["A1"].font = FONT_TITLE

    ws["A2"] = "15 criterios · 5 categorías · Escala de Saaty (1-9)"
    ws["A2"].font = FONT_SMALL

    # Tabla ranking
    row = 4
    headers = ["Ranking", "Portafolio", "Puntuación AHP", "Puntuación (%)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    for i, (_, r) in enumerate(ranking.iterrows()):
        row = 5 + i
        is_alt = i % 2 == 1
        ws.cell(row=row, column=1, value=int(r["ranking"]))
        ws.cell(row=row, column=2, value=r["portafolio"])
        ws.cell(row=row, column=3, value=r["score_ahp"])
        ws.cell(row=row, column=4, value=r["score_pct"])
        for c in range(1, 5):
            style_data_cell(ws, row, c, is_alt)
        ws.cell(row=row, column=3).number_format = "0.0000"
        ws.cell(row=row, column=4).number_format = "0.0\"%\""

        # Resaltar ganador
        if r["ranking"] == 1:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = FILL_GREEN
                ws.cell(row=row, column=c).font = Font(name="Arial", size=10, bold=True, color=GREEN)

    # Gráfico de barras
    n_portfolios = len(ranking)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"Puntuación AHP por portafolio ({profile})"
    chart.y_axis.title = "Puntuación (%)"
    chart.x_axis.title = "Portafolio"

    data_ref = Reference(ws, min_col=4, min_row=4, max_row=4 + n_portfolios)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=4 + n_portfolios)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 18
    chart.height = 12

    ws.add_chart(chart, "F4")

    # Pie chart
    pie = PieChart()
    pie.title = "Distribución de preferencia AHP"
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    pie.width = 14
    pie.height = 12

    ws.add_chart(pie, "F22")

    # Portafolio ganador destacado
    winner_row = 7 + n_portfolios
    ws.merge_cells(f"A{winner_row}:D{winner_row}")
    winner = ranking.iloc[0]
    ws.cell(row=winner_row, column=1,
            value=f"PORTAFOLIO RECOMENDADO: {winner['portafolio']} ({winner['score_pct']:.1f}%)")
    ws.cell(row=winner_row, column=1).font = FONT_WINNER

    auto_width(ws)


def create_consistency_sheet(wb: Workbook, ahp_report: Dict):
    """Hoja 4: Matrices AHP y consistencia."""
    ws = wb.create_sheet("4. Matrices AHP")

    ws.merge_cells("A1:D1")
    ws["A1"] = "Análisis de consistencia AHP"
    ws["A1"].font = FONT_TITLE

    row = 3
    ws.cell(row=row, column=1, value="Elemento").font = FONT_SUBHEADER
    ws.cell(row=row, column=2, value="Ratio Consistencia (CR)").font = FONT_SUBHEADER
    ws.cell(row=row, column=3, value="Consistente (CR < 0.10)").font = FONT_SUBHEADER
    style_header_row(ws, row, 3)

    row = 4
    ws.cell(row=row, column=1, value="Matriz de criterios")
    ws.cell(row=row, column=2, value=ahp_report.get("criteria_cr", 0))
    ws.cell(row=row, column=2).number_format = "0.0000"
    cr_val = ahp_report.get("criteria_cr", 0)
    ws.cell(row=row, column=3, value="SÍ" if cr_val < 0.10 else "NO")
    for c in range(1, 4):
        style_data_cell(ws, row, c)

    alt_crs = ahp_report.get("alt_crs", {})
    for i, (crit, cr) in enumerate(alt_crs.items()):
        row = 5 + i
        label = CRITERIA.get(crit, {}).get("label", crit)
        ws.cell(row=row, column=1, value=f"Alternativas ({label[:35]})")
        ws.cell(row=row, column=2, value=cr)
        ws.cell(row=row, column=2).number_format = "0.0000"
        ws.cell(row=row, column=3, value="SÍ" if cr < 0.10 else "NO")
        for c in range(1, 4):
            style_data_cell(ws, row, c, i % 2 == 1)
        if cr >= 0.10:
            ws.cell(row=row, column=3).fill = FILL_RED

    auto_width(ws)


# ─────────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL DE EXPORTACIÓN
# ─────────────────────────────────────────────────────────────────

def export_to_excel(
    profile: str,
    filters,
    portfolios: pd.DataFrame,
    ranking: pd.DataFrame,
    ahp_report: Dict,
    analysis: Optional[pd.DataFrame] = None,
    filename: str = "AHP_Portfolio_Result.xlsx",
    progress: bool = True,
) -> str:
    """
    Genera el archivo Excel completo con todos los resultados.

    Args:
        profile: nombre del perfil
        filters: UniverseFilters del perfil
        portfolios: DataFrame de portafolios candidatos
        ranking: DataFrame del ranking AHP
        ahp_report: dict del motor AHP (get_full_report())
        analysis: DataFrame del análisis individual (opcional)
        filename: nombre del archivo de salida

    Returns:
        Ruta del archivo generado.
    """
    if progress:
        print("\n" + "=" * 60)
        print("  MÓDULO 7 — EXPORTACIÓN A EXCEL")
        print(f"  Archivo: {filename}")
        print("=" * 60)

    wb = Workbook()
    # Eliminar la hoja por defecto
    wb.remove(wb.active)

    # Hoja 1: Perfil
    if progress:
        print("  Creando hoja 1: Perfil del inversor...")
    create_profile_sheet(wb, profile, filters)

    # Hoja 2: Análisis individual (si disponible)
    if analysis is not None and len(analysis) > 0:
        if progress:
            print("  Creando hoja 2: Análisis de acciones...")
        ws = wb.create_sheet("2. Análisis acciones")
        ws.merge_cells("A1:E1")
        ws["A1"] = "Análisis individual de acciones"
        ws["A1"].font = FONT_TITLE

        cols = [c for c in ["rentabilidad", "sharpe", "sortino", "alpha",
                           "volatilidad", "var_95", "max_drawdown", "beta", "cv"]
                if c in analysis.columns]

        headers = ["Ticker"] + [CRITERIA.get(c, {}).get("label", c)[:20] for c in cols]
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c, value=h)
        style_header_row(ws, 3, len(headers))

        for i, (ticker, row_data) in enumerate(analysis.head(30).iterrows()):
            r = 4 + i
            ws.cell(row=r, column=1, value=str(ticker))
            for j, col in enumerate(cols, 2):
                ws.cell(row=r, column=j, value=row_data.get(col, ""))
                style_data_cell(ws, r, j, i % 2 == 1)
                if col in ["rentabilidad", "volatilidad", "var_95", "max_drawdown"]:
                    ws.cell(row=r, column=j).number_format = "0.00%"
                else:
                    ws.cell(row=r, column=j).number_format = "0.00"
            style_data_cell(ws, r, 1, i % 2 == 1)
            ws.cell(row=r, column=1).alignment = LEFT

        auto_width(ws)

    # Hoja 3: Portafolios
    if progress:
        print("  Creando hoja 3: Portafolios candidatos...")
    create_portfolios_sheet(wb, portfolios)

    # Hoja 4: Matrices AHP
    if progress:
        print("  Creando hoja 4: Matrices AHP y consistencia...")
    create_consistency_sheet(wb, ahp_report)

    # Hoja 5: Ranking
    if progress:
        print("  Creando hoja 5: Ranking final...")
    create_ranking_sheet(wb, ranking, profile)

    # Guardar
    wb.save(filename)

    if progress:
        print(f"\n  Archivo guardado: {filename}")
        print(f"{'─' * 60}")

    return filename


# ─── DEMO ───
def demo():
    """Demo de exportación con datos ficticios."""
    from profiles import get_profile_config

    profile = "moderado"
    config = get_profile_config(profile)

    portfolios = pd.DataFrame({
        "nombre":       ["P1", "P2", "P3", "P4", "P5"],
        "rentabilidad": [0.086, 0.085, 0.056, 0.161, 0.104],
        "sharpe":       [0.48, 0.47, 0.15, 0.72, 0.55],
        "sortino":      [0.71, 0.70, 0.22, 1.05, 0.82],
        "alpha":        [0.02, 0.019, -0.01, 0.08, 0.04],
        "volatilidad":  [0.18, 0.18, 0.29, 0.22, 0.19],
        "var_95":       [0.049, 0.049, 0.078, 0.061, 0.054],
        "cvar_95":      [0.065, 0.064, 0.102, 0.081, 0.071],
        "max_drawdown": [0.15, 0.14, 0.31, 0.25, 0.18],
        "beta":         [0.85, 0.84, 1.10, 1.05, 0.90],
        "tracking_error": [0.08, 0.08, 0.14, 0.12, 0.09],
        "rentab_kp":    [0.025, 0.024, -0.015, 0.065, 0.035],
        "cv":           [2.09, 2.12, 5.18, 1.37, 1.83],
        "skewness":     [0.12, 0.10, -0.25, -0.15, 0.08],
        "corr_media":   [0.35, 0.36, 0.55, 0.42, 0.30],
        "div_geo":      [0.72, 0.70, 0.45, 0.55, 0.78],
    }).set_index("nombre")

    # Simular ranking AHP
    ranking = pd.DataFrame({
        "portafolio": ["P4", "P1", "P5", "P2", "P3"],
        "score_ahp":  [0.253, 0.239, 0.230, 0.246, 0.032],
        "score_pct":  [25.3, 23.9, 23.0, 24.6, 3.2],
        "ranking":    [1, 2, 3, 4, 5],
    })

    ahp_report = {
        "criteria_cr": 0.0078,
        "alt_crs": {c: 0.02 + i * 0.005 for i, c in enumerate(CRITERIA_ORDER)},
    }

    export_to_excel(
        profile=profile,
        filters=config["filters"],
        portfolios=portfolios,
        ranking=ranking,
        ahp_report=ahp_report,
        filename="AHP_Portfolio_Result_DEMO.xlsx",
    )


if __name__ == "__main__":
    demo()
